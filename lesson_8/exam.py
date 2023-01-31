# import deal
# from beartype import beartype
#
#
# @beartype
# @deal.pre(lambda _: num > 0)
# def test1(num : int) -> None:
#     print(num + 5)
#
#
#
# num = 0
# test1(num)


# ****************************************************

# file_data = open('text.txt', 'r')
#
# try:
#    #работа с файлами
#     pass
# finally:
#     file_data.close()
# for i in file_data.readlines():
#     print(i)

# with open('text.txt', 'r') as file:
#     print(file.read())

# file_data = open('text1.txt', 'w')
# file_data.write('Hello\nWorld')

# import json
#
# data = {'first_name': 'Nikolai',
#         'last_name': 'Nagornyi',
#         'age': 32,
#         'hobbies': ['tennis', 'skying'],
#         'children': [
#             {'first_name': 'Aleksandra',
#              'age': 3},
#             {'first_name': 'Oleg',
#              'age': 12},
#         ]
#
#         }
#
# with open('data.json', 'w') as file_data:
#     json.dump(data, file_data)

# import csv
#
# filename = 'csv_text.csv'
# with open(filename, 'r', encoding="utf-8") as file_data:
#     read_obj = csv.reader(file_data, delimiter=",")
#
#     for i in read_obj:
#         print(i)


import openpyxl

file_exl = openpyxl.load_workbook('exam.xlsx')
print(file_exl.sheetnames)

wb = file_exl.active
print(wb['A1'])
# for i in file_exl.sheetnames:
#     pass